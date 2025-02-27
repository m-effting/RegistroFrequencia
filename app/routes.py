# Importação das bibliotecas necessárias
from flask import render_template, request, redirect, url_for, jsonify, current_app, flash
import os
import csv
from datetime import datetime
from app import app
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Dicionário para armazenar as chamadas dos alunos
chamadas = {}

# Função para salvar as chamadas em um arquivo Excel
def salvar_chamada_excel(escola, turma):
    try:
        # Formatar a data para o nome do arquivo
        data = datetime.today().strftime('%Y-%m-%d')
        chave = f"{escola}_{turma}_{data}"

        # Usar o contexto do app para acessar o caminho do diretório de histórico
        with app.app_context():
            arquivo_excel = os.path.join(current_app.config['HISTORICO_ESCOLAR_FOLDER'], f"{chave}.xlsx")

        # Criar uma nova planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Chamada"

        # Adicionar cabeçalho na planilha
        ws.append(["Nome", "Presença", "Observação"])

        # Preencher as linhas com os dados dos alunos
        if chave in chamadas:
            for aluno_data in chamadas[chave]:
                ws.append([aluno_data['nome'], aluno_data['presenca'], aluno_data['observacao']])
                

        # Ajuste automático do tamanho das colunas
        for col in range(1, 4):
            column = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        # Encontrar o maior comprimento de valor para ajustar a largura
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)  # Ajustar a largura das colunas
            ws.column_dimensions[column].width = adjusted_width

        # Salvar o arquivo Excel
        wb.save(arquivo_excel)
        return arquivo_excel

    except Exception as e:
        # Em caso de erro, exibe uma mensagem flash e retorna None
        flash(f"Erro ao salvar chamada: {str(e)}", "error")
        return None

# Rota principal para exibir a página inicial
@app.route('/')
def index():
    turma_atual = request.args.get('turma', '')  # Obtém a turma da URL, se existir
    return render_template('index.html', chamadas=chamadas, turma_atual=turma_atual)

# Rota para adicionar um aluno à chamada
@app.route('/adicionar_aluno', methods=['POST'])
def adicionar_aluno():
    # Obtém os dados do formulário
    escola = request.form['escola']
    turma = request.form['turma']
    aluno = request.form['aluno']

    # Verifica se todos os campos foram preenchidos
    if not escola or not turma or not aluno:
        flash("Preencha todos os campos para adicionar um aluno.", "error")
        return redirect(url_for('index', turma=turma))

    # Formatar a data para chave única
    data = datetime.today().strftime('%Y-%m-%d')
    chave = f"{escola}_{turma}_{data}"

    # Se a chave não existe, cria uma nova lista de chamadas
    if chave not in chamadas:
        chamadas[chave] = []

    # Adicionar o aluno na lista de chamadas
    chamadas[chave].append({'nome': aluno, 'presenca': 'Presente', 'observacao': ''})

    # Salvar a chamada em Excel
    if salvar_chamada_excel(escola, turma):
        flash("Aluno adicionado com sucesso!", "success")
    else:
        flash("Erro ao salvar a chamada.", "error")

    return redirect(url_for('index', turma=turma))  # Redireciona mantendo a turma na URL

# Rota para alterar o status de presença de um aluno
@app.route('/alterar_presenca', methods=['POST'])
def alterar_presenca():
    escola = request.form.get('escola', 'Desconhecido')  # Obtém os dados de presença
    turma = request.form.get('turma', 'Desconhecida')
    aluno = request.form['aluno']
    status = request.form['status']

    # Formatar a data para chave única
    data = datetime.today().strftime('%Y-%m-%d')
    chave = f"{escola}_{turma}_{data}"

    # Verifica se a chave existe e altera o status de presença
    if chave in chamadas:
        for aluno_data in chamadas[chave]:
            if aluno_data['nome'] == aluno:
                aluno_data['presenca'] = status
                break

        # Salvar a alteração de presença em Excel
        if salvar_chamada_excel(escola, turma):
            return jsonify(success=True, message="Presença alterada com sucesso!")
        else:
            return jsonify(success=False, message="Erro ao salvar a alteração de presença.")
    
    return jsonify(success=False, message="Chamada não encontrada.")

# Rota para importar uma lista de presença a partir de um arquivo CSV
@app.route('/importar_chamada', methods=['POST'])
def importar_chamada():
    # Verifica se o arquivo foi enviado
    if 'file' not in request.files:
        flash("Nenhum arquivo selecionado.", "error")
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash("Nome de arquivo inválido.", "error")
        return redirect(url_for('index'))

    try:
        # Salvar o arquivo no diretório de uploads
        with app.app_context():
            filename = os.path.join(current_app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filename)

        # Dados da escola, turma e data para gerar a chave
        escola = request.form['escola']
        turma = request.form['turma']
        data = datetime.today().strftime('%Y-%m-%d')
        chave = f"{escola}_{turma}_{data}"

        # Se a chave não existe, cria uma nova lista de chamadas
        if chave not in chamadas:
            chamadas[chave] = []

        # Lê o arquivo CSV e adiciona os alunos na lista de chamadas
        with open(filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if row:
                    nome = row[0]  # Pega o nome do aluno da primeira coluna
                    chamadas[chave].append({'nome': nome, 'presenca': 'Presente', 'observacao': ''})

        # Salva a chamada importada em Excel
        if salvar_chamada_excel(escola, turma):
            flash("Chamada importada com sucesso!", "success")
        else:
            flash("Erro ao salvar a chamada importada.", "error")

    except Exception as e:
        # Exibe uma mensagem de erro em caso de falha ao importar
        flash(f"Erro ao importar chamada: {str(e)}", "error")

    return redirect(url_for('index'))
